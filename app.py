#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Gestión Misangeles - Aplicación Web
Aplicación web completa para gestión de inventario y ventas
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os
import json
from decimal import Decimal
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io

app = Flask(__name__)
app.config['SECRET_KEY'] = 'misangeles2025'
app.config['UPLOAD_FOLDER'] = 'static/img'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///misangeles.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ===== MODELOS DE BASE DE DATOS =====

class Configuracion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tasa_cambio = db.Column(db.Float, default=35.50)
    iva_porcentaje = db.Column(db.Float, default=30.0)
    fecha_dolar = db.Column(db.Date, default=date.today)
    fecha_programa = db.Column(db.Date, default=date.today)
    nombre_empresa = db.Column(db.String(100), default="Misangeles")
    
    def to_dict(self):
        return {
            'id': self.id,
            'tasa_cambio': self.tasa_cambio,
            'iva_porcentaje': self.iva_porcentaje,
            'fecha_dolar': self.fecha_dolar.strftime('%Y-%m-%d') if self.fecha_dolar else None,
            'fecha_programa': self.fecha_programa.strftime('%Y-%m-%d') if self.fecha_programa else None,
            'nombre_empresa': self.nombre_empresa
        }

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(20), unique=True, nullable=False)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    categoria = db.Column(db.String(50))
    precio_detal_usd = db.Column(db.Float, nullable=False)
    precio_mayor_usd = db.Column(db.Float, nullable=False)
    equivalencia = db.Column(db.Integer, default=1)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'categoria': self.categoria,
            'precio_detal_usd': self.precio_detal_usd,
            'precio_mayor_usd': self.precio_mayor_usd,
            'equivalencia': self.equivalencia,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None
        }

class Entrada(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=date.today)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    tipo_entrada = db.Column(db.String(10), nullable=False)  # 'Mayor' o 'Detal'
    cantidad = db.Column(db.Float, nullable=False)
    precio_unitario_usd = db.Column(db.Float, nullable=False)
    total_usd = db.Column(db.Float, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    producto = db.relationship('Producto', backref='entradas')
    
    def to_dict(self):
        return {
            'id': self.id,
            'fecha': self.fecha.strftime('%Y-%m-%d') if self.fecha else None,
            'producto_id': self.producto_id,
            'producto_codigo': self.producto.codigo if self.producto else None,
            'producto_nombre': self.producto.nombre if self.producto else None,
            'tipo_entrada': self.tipo_entrada,
            'cantidad': self.cantidad,
            'precio_unitario_usd': self.precio_unitario_usd,
            'total_usd': self.total_usd,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None
        }

class Venta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=date.today)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    tipo_venta = db.Column(db.String(10), nullable=False)  # 'Mayor' o 'Detal'
    cantidad_mayor = db.Column(db.Float, default=0)
    cantidad_detal = db.Column(db.Float, default=0)
    precio_unitario_usd = db.Column(db.Float, nullable=False)
    total_sin_iva_usd = db.Column(db.Float, nullable=False)
    iva_usd = db.Column(db.Float, nullable=False)
    total_con_iva_usd = db.Column(db.Float, nullable=False)
    total_bs = db.Column(db.Float, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    producto = db.relationship('Producto', backref='ventas')
    
    def to_dict(self):
        return {
            'id': self.id,
            'fecha': self.fecha.strftime('%Y-%m-%d') if self.fecha else None,
            'producto_id': self.producto_id,
            'producto_codigo': self.producto.codigo if self.producto else None,
            'producto_nombre': self.producto.nombre if self.producto else None,
            'tipo_venta': self.tipo_venta,
            'cantidad_mayor': self.cantidad_mayor,
            'cantidad_detal': self.cantidad_detal,
            'precio_unitario_usd': self.precio_unitario_usd,
            'total_sin_iva_usd': self.total_sin_iva_usd,
            'iva_usd': self.iva_usd,
            'total_con_iva_usd': self.total_con_iva_usd,
            'total_bs': self.total_bs,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None
        }

# ===== FUNCIONES AUXILIARES =====

def calcular_iva(monto, porcentaje_iva):
    """Calcula el IVA de un monto"""
    return monto * (porcentaje_iva / 100)

def convertir_usd_a_bs(monto_usd, tasa_cambio):
    """Convierte USD a BS"""
    return monto_usd * tasa_cambio

def obtener_configuracion():
    """Obtiene la configuración actual del sistema"""
    config = Configuracion.query.first()
    if not config:
        config = Configuracion()
        db.session.add(config)
        db.session.commit()
    return config

def calcular_stock_producto(producto_id):
    """Calcula el stock actual de un producto"""
    # Obtener entradas
    entradas_mayor = Entrada.query.filter_by(
        producto_id=producto_id, 
        tipo_entrada='Mayor'
    ).with_entities(db.func.sum(Entrada.cantidad)).scalar() or 0
    
    entradas_detal = Entrada.query.filter_by(
        producto_id=producto_id, 
        tipo_entrada='Detal'
    ).with_entities(db.func.sum(Entrada.cantidad)).scalar() or 0
    
    # Obtener ventas
    ventas_mayor = Venta.query.filter_by(
        producto_id=producto_id, 
        tipo_venta='Mayor'
    ).with_entities(db.func.sum(Venta.cantidad_mayor)).scalar() or 0
    
    ventas_detal = Venta.query.filter_by(
        producto_id=producto_id, 
        tipo_venta='Detal'
    ).with_entities(db.func.sum(Venta.cantidad_detal)).scalar() or 0
    
    # Calcular stock
    producto = Producto.query.get(producto_id)
    if not producto:
        return {'stock_mayor': 0, 'stock_unidades': 0, 'valor_usd': 0, 'valor_bs': 0}
    
    stock_mayor = entradas_mayor - ventas_mayor
    stock_unidades = (entradas_mayor * producto.equivalencia + entradas_detal) - ventas_detal
    
    # Calcular valores
    config = obtener_configuracion()
    valor_usd = (stock_mayor * producto.precio_mayor_usd) + (stock_unidades * producto.precio_detal_usd)
    valor_bs = valor_usd * config.tasa_cambio
    
    return {
        'stock_mayor': stock_mayor,
        'stock_unidades': stock_unidades,
        'valor_usd': round(valor_usd, 2),
        'valor_bs': round(valor_bs, 2)
    }

# ===== RUTAS DE LA APLICACIÓN =====

@app.route('/')
def dashboard():
    """Pantalla principal del sistema"""
    config = obtener_configuracion()
    
    # Obtener estadísticas del día
    hoy = date.today()
    ventas_hoy = Venta.query.filter_by(fecha=hoy).all()
    total_ventas_usd = sum(v.total_con_iva_usd for v in ventas_hoy)
    total_ventas_bs = sum(v.total_bs for v in ventas_hoy)
    
    # Obtener productos con stock
    productos = Producto.query.all()
    inventario = []
    total_inventario_usd = 0
    total_inventario_bs = 0
    
    for producto in productos:
        stock = calcular_stock_producto(producto.id)
        if stock['stock_mayor'] > 0 or stock['stock_unidades'] > 0:
            inventario.append({
                'producto': producto.to_dict(),
                'stock': stock
            })
            total_inventario_usd += stock['valor_usd']
            total_inventario_bs += stock['valor_bs']
    
    return render_template('dashboard.html',
                         config=config.to_dict(),
                         ventas_hoy=len(ventas_hoy),
                         total_ventas_usd=round(total_ventas_usd, 2),
                         total_ventas_bs=round(total_ventas_bs, 2),
                         inventario=inventario,
                         total_inventario_usd=round(total_inventario_usd, 2),
                         total_inventario_bs=round(total_inventario_bs, 2))

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    """Gestión de configuración del sistema"""
    if request.method == 'POST':
        config = obtener_configuracion()
        
        # Actualizar configuración
        config.tasa_cambio = float(request.form.get('tasa_cambio', 35.50))
        config.iva_porcentaje = float(request.form.get('iva_porcentaje', 30.0))
        config.nombre_empresa = request.form.get('nombre_empresa', 'Misangeles')
        
        db.session.commit()
        flash('Configuración actualizada correctamente', 'success')
        return redirect(url_for('configuracion'))
    
    config = obtener_configuracion()
    
    # Obtener URL del logo actual
    logo_url = get_logo_url()
    
    return render_template('configuracion.html', config=config, logo_url=logo_url)

@app.route('/productos')
def productos():
    """Lista de productos"""
    productos = Producto.query.all()
    return render_template('productos.html', productos=productos)

@app.route('/productos/nuevo', methods=['GET', 'POST'])
def nuevo_producto():
    """Crear nuevo producto"""
    if request.method == 'POST':
        producto = Producto(
            codigo=request.form['codigo'],
            nombre=request.form['nombre'],
            descripcion=request.form['descripcion'],
            categoria=request.form['categoria'],
            precio_detal_usd=float(request.form['precio_detal_usd']),
            precio_mayor_usd=float(request.form['precio_mayor_usd']),
            equivalencia=int(request.form['equivalencia'])
        )
        
        db.session.add(producto)
        db.session.commit()
        flash('Producto creado correctamente', 'success')
        return redirect(url_for('productos'))
    
    return render_template('nuevo_producto.html')

@app.route('/entradas')
def entradas():
    """Lista de entradas de inventario"""
    entradas = Entrada.query.order_by(Entrada.fecha.desc()).all()
    
    # Calcular estadísticas
    hoy = date.today()
    entradas_hoy = Entrada.query.filter_by(fecha=hoy).count()
    total_usd = sum(e.total_usd for e in entradas)
    config = obtener_configuracion()
    total_bs = total_usd * config.tasa_cambio
    
    return render_template('entradas.html', 
                         entradas=entradas,
                         entradas_hoy=entradas_hoy,
                         total_usd=total_usd,
                         total_bs=total_bs)

@app.route('/entradas/nueva', methods=['GET', 'POST'])
def nueva_entrada():
    """Crear nueva entrada de inventario"""
    if request.method == 'POST':
        producto = Producto.query.filter_by(codigo=request.form['codigo_producto']).first()
        if not producto:
            flash('Producto no encontrado', 'error')
            return redirect(url_for('nueva_entrada'))
        
        tipo_entrada = request.form['tipo_entrada']
        cantidad = float(request.form['cantidad'])
        fecha_entrada = datetime.strptime(request.form['fecha'], '%Y-%m-%d').date()
        
        # Determinar precio según tipo
        if tipo_entrada == 'Mayor':
            precio_unitario = producto.precio_mayor_usd
        else:
            precio_unitario = producto.precio_detal_usd
        
        total_usd = cantidad * precio_unitario
        
        entrada = Entrada(
            producto_id=producto.id,
            tipo_entrada=tipo_entrada,
            cantidad=cantidad,
            precio_unitario_usd=precio_unitario,
            total_usd=total_usd,
            fecha=fecha_entrada
        )
        
        db.session.add(entrada)
        db.session.commit()
        flash('Entrada registrada correctamente', 'success')
        return redirect(url_for('entradas'))
    
    productos = Producto.query.all()
    config = obtener_configuracion()
    today = date.today()
    
    return render_template('nueva_entrada.html', 
                         productos=productos,
                         config=config.to_dict(),
                         today=today)

@app.route('/ventas')
def ventas():
    """Lista de ventas"""
    ventas = Venta.query.order_by(Venta.fecha.desc()).all()
    
    # Calcular estadísticas
    hoy = date.today()
    ventas_hoy = Venta.query.filter_by(fecha=hoy).count()
    total_usd = sum(v.total_con_iva_usd for v in ventas)
    total_bs = sum(v.total_bs for v in ventas)
    
    return render_template('ventas.html', 
                         ventas=ventas,
                         ventas_hoy=ventas_hoy,
                         total_usd=total_usd,
                         total_bs=total_bs)

@app.route('/ventas/nueva', methods=['GET', 'POST'])
def nueva_venta():
    """Registrar nueva venta"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            codigo = request.form['codigo']
            fecha_venta = datetime.strptime(request.form['fecha'], '%Y-%m-%d').date()
            tipo_venta = request.form['tipo_venta']
            cantidad = float(request.form['cantidad'])
            precio_unitario = float(request.form['precio_unitario'])
            observaciones = request.form.get('observaciones', '')
            
            # Buscar producto
            producto = Producto.query.filter_by(codigo=codigo).first()
            if not producto:
                flash('Producto no encontrado', 'danger')
                return redirect(url_for('nueva_venta'))
            
            # Calcular totales
            subtotal_usd = cantidad * precio_unitario
            config = obtener_configuracion()
            iva_usd = subtotal_usd * (config.iva_porcentaje / 100)
            total_con_iva_usd = subtotal_usd + iva_usd
            total_bs = total_con_iva_usd * config.tasa_cambio
            
            # Crear nueva venta
            nueva_venta = Venta(
                fecha=fecha_venta,
                producto_id=producto.id,
                tipo_venta=tipo_venta,
                cantidad_mayor=cantidad if tipo_venta == 'Mayor' else 0,
                cantidad_detal=cantidad if tipo_venta == 'Detal' else 0,
                precio_unitario_usd=precio_unitario,
                total_sin_iva_usd=subtotal_usd,
                iva_usd=iva_usd,
                total_con_iva_usd=total_con_iva_usd,
                total_bs=total_bs
            )
            
            db.session.add(nueva_venta)
            db.session.commit()
            
            flash('Venta registrada exitosamente', 'success')
            return redirect(url_for('ventas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar la venta: {str(e)}', 'danger')
            return redirect(url_for('nueva_venta'))
    
    productos = Producto.query.all()
    config = obtener_configuracion()
    today = date.today()
    
    return render_template('nueva_venta.html', 
                         productos=productos,
                         config=config.to_dict(),
                         today=today)

@app.route('/inventario')
def inventario():
    """Vista del inventario actual"""
    productos = Producto.query.all()
    inventario = []
    total_inventario_usd = 0
    total_inventario_bs = 0
    
    for producto in productos:
        stock = calcular_stock_producto(producto.id)
        if stock['stock_mayor'] > 0 or stock['stock_unidades'] > 0:
            inventario.append({
                'producto': producto.to_dict(),
                'stock': stock
            })
            total_inventario_usd += stock['valor_usd']
            total_inventario_bs += stock['valor_bs']
    
    return render_template('inventario.html', 
                         inventario=inventario,
                         total_inventario_usd=total_inventario_usd,
                         total_inventario_bs=total_inventario_bs)

@app.route('/reportes')
def reportes():
    """Página de reportes"""
    # Obtener estadísticas básicas
    total_productos = Producto.query.count()
    
    # Obtener ventas del mes actual
    hoy = date.today()
    inicio_mes = date(hoy.year, hoy.month, 1)
    ventas_mes = Venta.query.filter(Venta.fecha >= inicio_mes).count()
    
    # Calcular totales de ventas del mes
    ventas_mes_data = Venta.query.filter(Venta.fecha >= inicio_mes).all()
    total_ventas_usd = sum(v.total_con_iva_usd for v in ventas_mes_data)
    total_ventas_bs = sum(v.total_bs for v in ventas_mes_data)
    
    return render_template('reportes.html',
                         total_productos=total_productos,
                         ventas_mes=ventas_mes,
                         total_ventas_usd=total_ventas_usd,
                         total_ventas_bs=total_ventas_bs)

@app.route('/exportar/excel')
def exportar_excel():
    """Exportar inventario a Excel"""
    productos = Producto.query.all()
    inventario = []
    
    for producto in productos:
        stock = calcular_stock_producto(producto.id)
        inventario.append({
            'Código': producto.codigo,
            'Nombre': producto.nombre,
            'Descripción': producto.descripcion,
            'Categoría': producto.categoria,
            'Stock Mayor': stock['stock_mayor'],
            'Stock Unidades': stock['stock_unidades'],
            'Valor USD': stock['valor_usd'],
            'Valor BS': stock['valor_bs']
        })
    
    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Encabezados
    if inventario:
        headers = list(inventario[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        for row, item in enumerate(inventario, 2):
            for col, value in enumerate(item.values(), 1):
                ws.cell(row=row, column=col, value=value)
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_misangeles_{date.today()}.xlsx'
    )

@app.route('/api/productos/<codigo>')
def api_producto_por_codigo(codigo):
    """API para buscar producto por código"""
    producto = Producto.query.filter_by(codigo=codigo).first()
    if producto:
        return jsonify(producto.to_dict())
    return jsonify({'error': 'Producto no encontrado'}), 404

@app.route('/api/inventario')
def api_inventario():
    """API para obtener inventario completo"""
    productos = Producto.query.all()
    inventario = []
    
    for producto in productos:
        stock = calcular_stock_producto(producto.id)
        inventario.append({
            'producto': producto.to_dict(),
            'stock': stock
        })
    
    return jsonify(inventario)

@app.route('/subir_logo', methods=['POST'])
def subir_logo():
    try:
        if 'logo' not in request.files:
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        file = request.files['logo']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        if file and allowed_file(file.filename):
            # Guardar como logo.png (reemplaza el anterior)
            filename = "logo.png"
            filepath = os.path.join(app.root_path, 'static', 'img', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            # Guardar archivo
            file.save(filepath)
            
            # Crear URL para el archivo
            logo_url = url_for('static', filename=f'img/{filename}')
            
            return jsonify({
                'success': True, 
                'message': 'Logo subido correctamente',
                'logo_url': logo_url
            })
        else:
            return jsonify({'success': False, 'message': 'Formato de archivo no permitido'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al subir logo: {str(e)}'})

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_logo_url():
    """Obtiene la URL del logo actual del sistema"""
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        return url_for('static', filename='img/logo.png')
    return url_for('static', filename='img/default-logo.png')

# ===== INICIALIZACIÓN =====

def inicializar_base_datos():
    """Inicializa la base de datos con datos de ejemplo"""
    with app.app_context():
        db.create_all()
        
        # Crear configuración por defecto
        if not Configuracion.query.first():
            config = Configuracion()
            db.session.add(config)
            db.session.commit()
        
        # Crear productos de ejemplo
        if not Producto.query.first():
            productos_ejemplo = [
                Producto(
                    codigo='H001',
                    nombre='Huevos',
                    descripcion='Huevos de gallina frescos',
                    categoria='Proteínas',
                    precio_detal_usd=0.15,
                    precio_mayor_usd=0.12,
                    equivalencia=30
                ),
                Producto(
                    codigo='A001',
                    nombre='Arroz',
                    descripcion='Arroz blanco premium',
                    categoria='Granos',
                    precio_detal_usd=2.50,
                    precio_mayor_usd=2.00,
                    equivalencia=25
                ),
                Producto(
                    codigo='AC001',
                    nombre='Aceite',
                    descripcion='Aceite de cocina vegetal',
                    categoria='Aceites',
                    precio_detal_usd=3.00,
                    precio_mayor_usd=2.40,
                    equivalencia=12
                )
            ]
            
            for producto in productos_ejemplo:
                db.session.add(producto)
            
            db.session.commit()

if __name__ == '__main__':
    inicializar_base_datos()
    print("🚀 Sistema de Gestión Mis Angeles iniciado")
    print("📱 Accede a: http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
